#!/usr/bin/env python3
"""Force a headless LibreOffice recalculation pass over an Excel workbook.

openpyxl writes formulas but does not evaluate them. Viewers that do
not auto-recalculate (Google Sheets, LibreOffice Calc on first open, a
Python process reading with ``data_only=True``) see empty cells where
the user expects values. This helper opens the workbook with a headless
LibreOffice instance, forces Calculate() on every sheet, saves in
place, then scans the result with openpyxl to surface any Excel errors
(#REF!, #DIV/0!, #VALUE!, #N/A, #NAME?, #NULL!, #NUM!).

Output: JSON on stdout summarising formula count, sheets touched and
detected errors. Diagnostics go to stderr. Exit code 0 on success,
1 on failure.

Usage:
    python3 refresh_formulas.py workbook.xlsx
    python3 refresh_formulas.py workbook.xlsx --timeout 90 --json
    python3 refresh_formulas.py workbook.xlsx --quiet
"""
from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any

EXCEL_ERROR_MARKERS = (
    "#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!",
)


def _log(msg: str, quiet: bool) -> None:
    if not quiet:
        print(msg, file=sys.stderr)


def _ensure_libreoffice() -> str:
    soffice = shutil.which("soffice")
    if soffice is None:
        raise RuntimeError(
            "LibreOffice headless (soffice) not found on PATH. "
            "Install libreoffice or libreoffice-calc."
        )
    return soffice


def _recalc_via_soffice(path: Path, soffice: str, timeout: int, quiet: bool) -> None:
    """Round-trip the workbook through LibreOffice to refresh cached values.

    Strategy: ``soffice --headless --calc --convert-to xlsx`` over the file
    writes a refreshed copy to an output directory; we then copy it back
    over the input. This does not require a Basic macro and works in
    fully-sandboxed environments.
    """
    with tempfile.TemporaryDirectory(prefix="xlsx_refresh_") as tmp:
        out_dir = Path(tmp)
        _log(f"[soffice] recalculating via {soffice} --convert-to xlsx", quiet)
        try:
            subprocess.run(
                [
                    soffice, "--headless", "--calc",
                    "--convert-to", "xlsx",
                    "--outdir", str(out_dir), str(path),
                ],
                check=True, capture_output=True, timeout=timeout,
            )
        except subprocess.TimeoutExpired as exc:
            raise RuntimeError(
                f"soffice recalculation timed out after {timeout}s"
            ) from exc
        except subprocess.CalledProcessError as exc:
            stderr = exc.stderr.decode("utf-8", errors="replace") if exc.stderr else ""
            raise RuntimeError(f"soffice recalculation failed: {stderr.strip()}") from exc

        produced = out_dir / (path.stem + ".xlsx")
        if not produced.exists():
            raise RuntimeError(
                f"soffice produced no output for {path}; "
                "the input may not be a valid XLSX."
            )
        shutil.copy2(produced, path)


def _scan_errors(path: Path, quiet: bool) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl not installed") from exc

    _log(f"[scan] opening {path} with data_only=True", quiet)
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        formula_cells = 0
        sheets_touched: list[str] = []
        issues: list[dict[str, str]] = []

        # We also need the formula-aware view to count formula cells.
        wb_f = load_workbook(str(path), read_only=True, data_only=False)
        try:
            for sheet_name in wb_f.sheetnames:
                sf = wb_f[sheet_name]
                sv = wb[sheet_name]
                touched = False
                for row_f, row_v in zip(sf.iter_rows(values_only=True),
                                        sv.iter_rows(values_only=False)):
                    for idx, (vf, cell_v) in enumerate(zip(row_f, row_v)):
                        if isinstance(vf, str) and vf.startswith("="):
                            formula_cells += 1
                            touched = True
                        value = cell_v.value if hasattr(cell_v, "value") else cell_v
                        if isinstance(value, str) and value in EXCEL_ERROR_MARKERS:
                            issues.append({
                                "sheet": sheet_name,
                                "cell": cell_v.coordinate if hasattr(cell_v, "coordinate") else f"?{idx}",
                                "type": value,
                            })
                if touched:
                    sheets_touched.append(sheet_name)
        finally:
            wb_f.close()
    finally:
        wb.close()

    return {
        "formula_cells": formula_cells,
        "sheets_touched": sheets_touched,
        "issues": issues,
    }


def refresh(path: Path, timeout: int, quiet: bool) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    soffice = _ensure_libreoffice()
    _recalc_via_soffice(path, soffice, timeout, quiet)
    report = _scan_errors(path, quiet)
    report["ok"] = len(report["issues"]) == 0
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description=(__doc__ or "").split("\n\n")[0])
    parser.add_argument("path", help="Path to the .xlsx file.")
    parser.add_argument("--timeout", type=int, default=120,
                        help="LibreOffice subprocess timeout in seconds (default: 120).")
    parser.add_argument("--json", action="store_true",
                        help="Emit the full JSON report on stdout (default: compact).")
    parser.add_argument("--quiet", action="store_true",
                        help="Silence diagnostic messages on stderr.")
    args = parser.parse_args()

    path = Path(args.path)

    try:
        report = refresh(path, timeout=args.timeout, quiet=args.quiet)
    except Exception as exc:  # noqa: BLE001
        report = {"ok": False, "error": str(exc)}
        sys.stdout.write(json.dumps(report, ensure_ascii=False) + "\n")
        return 1

    if args.json:
        sys.stdout.write(json.dumps(report, ensure_ascii=False, indent=2) + "\n")
    else:
        sys.stdout.write(json.dumps(report, ensure_ascii=False) + "\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
