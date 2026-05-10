#!/usr/bin/env python3
"""Quick extraction from an Excel workbook.

Returns structured output on stdout (Markdown tables by default, CSV or JSON
on request). Diagnostics go to stderr. Exit code 0 on success, 1 on failure.
The engine chain is: openpyxl (read-only) -> pandas (openpyxl backend) ->
raw-XML walk. The first one that produces non-empty rows for the requested
sheet wins.

Usage:
    python3 quick_extract.py workbook.xlsx
    python3 quick_extract.py workbook.xlsx --sheet Summary
    python3 quick_extract.py workbook.xlsx --sheet all --format csv
    python3 quick_extract.py workbook.xlsx --format json --max-rows 500
    cat workbook.xlsx | python3 quick_extract.py -
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Iterable

SUPPORTED_ENGINES = ("auto", "openpyxl", "pandas", "unzip")
SUPPORTED_FORMATS = ("md", "csv", "json")


def _log(msg: str) -> None:
    print(msg, file=sys.stderr)


def _read_stdin_to_tempfile() -> Path:
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(sys.stdin.buffer.read())
    tmp.close()
    return Path(tmp.name)


def _convert_legacy_xls(path: Path) -> Path:
    """Convert a legacy .xls or .xlsb to .xlsx via soffice. Returns new path."""
    if shutil.which("soffice") is None:
        raise RuntimeError(
            "File looks like a legacy .xls / .xlsb but 'soffice' (LibreOffice) "
            "is not installed. Install libreoffice or libreoffice-calc."
        )
    out_dir = Path(tempfile.mkdtemp(prefix="xlsx_conv_"))
    subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx",
         "--outdir", str(out_dir), str(path)],
        check=True, capture_output=True,
    )
    converted = out_dir / (path.stem + ".xlsx")
    if not converted.exists():
        raise RuntimeError(f"LibreOffice produced no output for {path}")
    return converted


def _is_xlsx_zip(path: Path) -> bool:
    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            return any(n.startswith("xl/") for n in names)
    except zipfile.BadZipFile:
        return False


def _trim_trailing_empty(rows: list[list[Any]]) -> list[list[Any]]:
    while rows and all(c is None or (isinstance(c, str) and c == "") for c in rows[-1]):
        rows.pop()
    return rows


def _rows_with_openpyxl(
    path: Path, sheet_target: str, max_rows: int | None,
) -> dict[str, list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl not installed") from exc

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
        target_sheets: list[str]
        if sheet_target == "all":
            target_sheets = list(sheet_names)
        elif sheet_target == "__active__":
            active = wb.active
            active_title = active.title if active is not None else sheet_names[0]
            target_sheets = [active_title]
        else:
            if sheet_target not in sheet_names:
                raise RuntimeError(
                    f"sheet {sheet_target!r} not in workbook; available: {sheet_names}"
                )
            target_sheets = [sheet_target]

        out: dict[str, list[list[Any]]] = {}
        for name in target_sheets:
            ws = wb[name]
            rows: list[list[Any]] = []
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if max_rows is not None and idx >= max_rows:
                    break
                rows.append(list(row))
            out[name] = _trim_trailing_empty(rows)
        return out
    finally:
        wb.close()


def _rows_with_pandas(
    path: Path, sheet_target: str, max_rows: int | None,
) -> dict[str, list[list[Any]]]:
    try:
        import pandas as pd
    except ImportError as exc:
        raise RuntimeError("pandas not installed") from exc

    sheet_name: Any
    if sheet_target == "all":
        sheet_name = None  # pandas: dict of all sheets
    elif sheet_target == "__active__":
        sheet_name = 0
    else:
        sheet_name = sheet_target

    frames = pd.read_excel(str(path), sheet_name=sheet_name, nrows=max_rows, header=None)
    if not isinstance(frames, dict):
        frames = {"Sheet1" if sheet_target == "__active__" else sheet_target: frames}

    out: dict[str, list[list[Any]]] = {}
    for name, df in frames.items():
        rows = [[_coerce_scalar(v) for v in row] for row in df.itertuples(index=False, name=None)]
        out[str(name)] = _trim_trailing_empty(rows)
    return out


def _coerce_scalar(v: Any) -> Any:
    # pandas/openpyxl return a mix of types; normalise NaN/NaT to None for portability.
    import math
    if v is None:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    to_pydt = getattr(v, "to_pydatetime", None)
    if callable(to_pydt):
        try:
            dt = to_pydt()
            iso = getattr(dt, "isoformat", None)
            if callable(iso):
                return iso()
        except Exception:
            return v
    return v


def _rows_with_unzip(
    path: Path, sheet_target: str, max_rows: int | None,
) -> dict[str, list[list[Any]]]:
    from lxml import etree  # type: ignore[import-untyped]

    NS_S = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns = {"s": NS_S}

    with zipfile.ZipFile(path) as z:
        # Resolve sheet names and relationship to their XML parts
        with z.open("xl/workbook.xml") as f:
            wb_tree = etree.parse(f)
        sheets = [
            (el.get("name"), el.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"))
            for el in wb_tree.getroot().find("s:sheets", ns)
        ]
        with z.open("xl/_rels/workbook.xml.rels") as f:
            rel_tree = etree.parse(f)
        rel_target = {
            r.get("Id"): r.get("Target")
            for r in rel_tree.getroot()
        }
        # Shared strings (optional)
        strings: list[str] = []
        if "xl/sharedStrings.xml" in z.namelist():
            with z.open("xl/sharedStrings.xml") as f:
                ss_tree = etree.parse(f)
            for si in ss_tree.getroot().findall("s:si", ns):
                text = "".join(t.text or "" for t in si.iter(f"{{{NS_S}}}t"))
                strings.append(text)

        sheet_map = {name: rel_target[rid] for name, rid in sheets}
        if sheet_target == "all":
            target_names = list(sheet_map.keys())
        elif sheet_target == "__active__":
            target_names = list(sheet_map.keys())[:1]
        else:
            if sheet_target not in sheet_map:
                raise RuntimeError(
                    f"sheet {sheet_target!r} not in workbook; available: {list(sheet_map)}"
                )
            target_names = [sheet_target]

        out: dict[str, list[list[Any]]] = {}
        for name in target_names:
            target = sheet_map[name]
            part = target if target.startswith("xl/") else "xl/" + target.lstrip("/")
            if part not in z.namelist():
                out[name] = []
                continue
            with z.open(part) as f:
                sheet_tree = etree.parse(f)
            rows_out: list[list[Any]] = []
            for idx, row in enumerate(sheet_tree.getroot().iter(f"{{{NS_S}}}row")):
                if max_rows is not None and idx >= max_rows:
                    break
                row_values: list[Any] = []
                for c in row.findall("s:c", ns):
                    t = c.get("t")
                    v = c.find("s:v", ns)
                    if v is None or v.text is None:
                        row_values.append(None)
                        continue
                    raw = v.text
                    if t == "s":
                        try:
                            row_values.append(strings[int(raw)])
                        except (ValueError, IndexError):
                            row_values.append(raw)
                    elif t == "b":
                        row_values.append(raw == "1")
                    elif t in ("str", "inlineStr"):
                        row_values.append(raw)
                    else:
                        # numeric: try int then float
                        try:
                            as_int = int(raw)
                            if str(as_int) == raw:
                                row_values.append(as_int)
                            else:
                                row_values.append(float(raw))
                        except ValueError:
                            try:
                                row_values.append(float(raw))
                            except ValueError:
                                row_values.append(raw)
                rows_out.append(row_values)
            out[name] = _trim_trailing_empty(rows_out)
        return out


def _format_md(sheet_data: dict[str, list[list[Any]]]) -> str:
    parts: list[str] = []
    for name, rows in sheet_data.items():
        parts.append(f"## {name}")
        if not rows:
            parts.append("_(empty)_")
            continue
        max_cols = max(len(r) for r in rows)
        padded = [[_to_cell(c) for c in (r + [None] * (max_cols - len(r)))] for r in rows]
        header = padded[0]
        body = padded[1:]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join(["---"] * max_cols) + " |")
        for r in body:
            parts.append("| " + " | ".join(r) + " |")
    return "\n\n".join(parts)


def _to_cell(v: Any) -> str:
    if v is None:
        return " "
    if isinstance(v, bool):
        return "true" if v else "false"
    s = str(v).replace("|", "\\|").replace("\n", " ")
    return s or " "


def _format_csv(sheet_data: dict[str, list[list[Any]]]) -> str:
    out = io.StringIO()
    multi = len(sheet_data) > 1
    for name, rows in sheet_data.items():
        if multi:
            out.write(f"# sheet: {name}\n")
        writer = csv.writer(out)
        for r in rows:
            writer.writerow(["" if v is None else v for v in r])
        if multi:
            out.write("\n")
    return out.getvalue().rstrip("\n")


def _format_json(sheet_data: dict[str, list[list[Any]]]) -> str:
    # Emit one object per row using row 0 as the header, per sheet
    payload: dict[str, list[dict[str, Any]]] = {}
    for name, rows in sheet_data.items():
        if not rows:
            payload[name] = []
            continue
        header = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        payload[name] = [
            {header[i] if i < len(header) else f"col_{i}": r[i] if i < len(r) else None
             for i in range(len(header))}
            for r in rows[1:]
        ]
    return json.dumps(payload, ensure_ascii=False, indent=2, default=str)


def extract(
    path: Path, engine: str, sheet_target: str, max_rows: int | None,
) -> dict[str, list[list[Any]]]:
    chain: Iterable[str] = (
        ("openpyxl", "pandas", "unzip") if engine == "auto" else (engine,)
    )
    errors: list[str] = []
    for name in chain:
        try:
            if name == "openpyxl":
                data = _rows_with_openpyxl(path, sheet_target, max_rows)
            elif name == "pandas":
                data = _rows_with_pandas(path, sheet_target, max_rows)
            elif name == "unzip":
                data = _rows_with_unzip(path, sheet_target, max_rows)
            else:
                raise RuntimeError(f"unknown engine {name!r}")
        except Exception as exc:  # noqa: BLE001
            _log(f"[{name}] failed: {exc}")
            errors.append(f"{name}: {exc}")
            continue
        non_empty = any(rows for rows in data.values())
        if non_empty:
            total = sum(len(rs) for rs in data.values())
            _log(f"[{name}] produced {len(data)} sheet(s), {total} rows total")
            return data
        _log(f"[{name}] produced empty output, trying next engine")
    raise RuntimeError(
        "All engines failed or produced empty output. Errors: " + "; ".join(errors)
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=(__doc__ or "").split("\n\n")[0])
    parser.add_argument("path", help="Path to the .xlsx / .xlsm / .xls file, or '-' for stdin.")
    parser.add_argument("--sheet", default="__active__",
                        help="Sheet name, 'all', or omit to use the active sheet.")
    parser.add_argument("--format", choices=SUPPORTED_FORMATS, default="md",
                        help="Output format (default: md).")
    parser.add_argument("--max-rows", type=int, default=None,
                        help="Cap rows per sheet.")
    parser.add_argument("--engine", choices=SUPPORTED_ENGINES, default="auto",
                        help="Force a specific engine (default: auto fallback).")
    args = parser.parse_args()

    if args.path == "-":
        path = _read_stdin_to_tempfile()
    else:
        path = Path(args.path)
        if not path.exists():
            _log(f"error: file not found: {path}")
            return 1

    try:
        if not _is_xlsx_zip(path):
            _log("input is not a .xlsx ZIP; attempting legacy .xls / .xlsb conversion")
            path = _convert_legacy_xls(path)
    except Exception as exc:  # noqa: BLE001
        _log(f"error: cannot read input: {exc}")
        return 1

    try:
        data = extract(path, engine=args.engine, sheet_target=args.sheet,
                       max_rows=args.max_rows)
    except Exception as exc:  # noqa: BLE001
        _log(f"error: {exc}")
        return 1

    if args.format == "md":
        out = _format_md(data)
    elif args.format == "csv":
        out = _format_csv(data)
    else:
        out = _format_json(data)

    sys.stdout.write(out)
    sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
